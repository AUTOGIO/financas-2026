#!/usr/bin/env python3
"""
analytics_engine.py — financas-2026
────────────────────────────────────────────────────────────────
Ported from giovannini-finance/scripts/compute_insights.py
(z-score anomaly detection + rolling/linear-trend projection).

ADAPTATION NOTE (read before trusting the numbers):
financas-2026's data model does not have what the original script
assumed:
  - No income tracking exists anywhere in this repo -> there is no
    cash-flow (income - expense) projection here, only an EXPENSE
    projection off weekly card spend.
  - Category detail only exists for the "AI/subscription" subset in
    the Transactions sheet (72 rows) — NOT the full spend ledger
    (Banks/MP Faturas/International are untouched). Anomaly
    detection below is scoped to that subset. It will not see
    anomalies in, say, Mercado Pago general spend.
If those two gaps become a real problem, the fix is building a
proper month x category matrix off the full ledger — not this
script. See README note in this file's output for a reminder.

Reads  : data/financas2026-DataEntry.xlsx
           - '💳 Transactions' sheet (AI/subscription card charges)
           - '📅 Weekly' sheet (weekly card spend, all categories)
Outputs: data/insights.json  (new file — does not touch DataEntry.xlsx
         or the dashboard HTML's SYNC block)

Usage:
    python3 scripts/analytics_engine.py
"""

from __future__ import annotations

import json
import statistics
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "Missing dependency: openpyxl. Install with:\n"
        "  python3 -m venv .venv && source .venv/bin/activate && pip install -r requirements.txt\n"
        "or, if you insist on system Python:\n"
        "  pip3 install --user openpyxl"
    )

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data" / "financas2026-DataEntry.xlsx"
OUTPUT_PATH = ROOT / "data" / "insights.json"
STRAY_ROOT_WORKBOOK = ROOT / "financas2026-DataEntry.xlsx"

Z_THRESHOLD = 1.5
ANOMALY_WINDOW = 3  # months of baseline


# ── validation ────────────────────────────────────────────────────────────
def require_workbook() -> None:
    # Guard against a stray copy at the repo root shadowing the tracked one.
    # AGENTS.md forbids root-level workbooks; if it reappears the operator is
    # almost certainly editing it while this script silently reads the old
    # data/ copy — see AUDIT-003 in REPOSITORY_AUDIT.md.
    if STRAY_ROOT_WORKBOOK.exists():
        sys.exit(
            f"Refusing to run: a workbook exists at the repo root:\n"
            f"  {STRAY_ROOT_WORKBOOK}\n"
            f"That shadows the canonical location at:\n"
            f"  {WORKBOOK_PATH}\n"
            "Move or delete the root copy (the operator likely edited the wrong file)."
        )
    if not WORKBOOK_PATH.exists():
        sys.exit(f"Not found: {WORKBOOK_PATH}\nExpected the data-entry workbook at this path.")


def require_sheet(wb, name: str):
    if name not in wb.sheetnames:
        sys.exit(
            f"Sheet '{name}' not found in {WORKBOOK_PATH.name}.\n"
            f"Available sheets: {wb.sheetnames}\n"
            "The workbook layout may have changed — update this script's "
            "sheet name / column indices before trusting the output."
        )


# ── load: Transactions (AI/subscription subset) ─────────────────────────────
def load_transactions(wb) -> list[dict]:
    require_sheet(wb, "💳 Transactions")
    ws = wb["💳 Transactions"]
    rows = list(ws.iter_rows(values_only=True))
    # rows[0]=title, rows[1]=subtitle, rows[2]=header, rows[3:]=data
    out = []
    for r in rows[3:]:
        if not r or r[3] is None or r[5] is None or r[6] is None:
            continue
        tx_date = r[3]
        category = r[5]
        amount = abs(float(r[6]))
        if not isinstance(tx_date, datetime):
            continue
        out.append({
            "date": tx_date.date().isoformat(),
            "month": tx_date.strftime("%Y-%m"),
            "merchant": r[4],
            "category": category,
            "amount": amount,
            "decision": r[8],
        })
    return out


# ── load: Weekly (all-category card spend) ───────────────────────────────────
def load_weekly(wb) -> list[dict]:
    require_sheet(wb, "📅 Weekly")
    ws = wb["📅 Weekly"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[3:]:
        if not r or r[1] is None or r[4] is None:
            continue
        out.append({
            "week": r[1],
            "start": r[2].date().isoformat() if isinstance(r[2], datetime) else None,
            "spend": abs(float(r[4])),
            "txs": r[5],
        })
    return out


# ── monthly x category aggregation (Transactions subset only) ───────────────
def build_monthly_category_agg(txs: list[dict]) -> dict[str, dict[str, float]]:
    months: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for t in txs:
        months[t["month"]][t["category"]] += t["amount"]
    return {m: dict(cats) for m, cats in sorted(months.items())}


# ── z-score anomaly detection (ported from giovannini-finance) ──────────────
def detect_anomalies(
    monthly_cat: dict[str, dict[str, float]],
    window: int = ANOMALY_WINDOW,
    z_threshold: float = Z_THRESHOLD,
) -> list[dict]:
    sorted_months = sorted(monthly_cat.keys())
    if len(sorted_months) < window + 1:
        return []

    cat_series: dict[str, dict[str, float]] = defaultdict(dict)
    for m in sorted_months:
        for cat, amt in monthly_cat[m].items():
            cat_series[cat][m] = amt

    recent_month = sorted_months[-1]
    baseline_months = sorted_months[-(window + 1):-1]
    anomalies: list[dict] = []

    for cat, series in cat_series.items():
        baseline_vals = [series.get(m, 0.0) for m in baseline_months]
        current_val = series.get(recent_month, 0.0)
        if current_val == 0 or len(baseline_vals) < 2:
            continue
        mean = statistics.mean(baseline_vals)
        stdev = statistics.stdev(baseline_vals)
        if stdev < 1.0:
            continue
        z = (current_val - mean) / stdev
        if abs(z) >= z_threshold:
            severity = "high" if abs(z) >= 2.5 else "medium" if abs(z) >= 2.0 else "low"
            anomalies.append({
                "category": cat,
                "month": recent_month,
                "current": round(current_val, 2),
                "baseline_mean": round(mean, 2),
                "z_score": round(z, 2),
                "direction": "acima" if z > 0 else "abaixo",
                "severity": severity,
            })
    return sorted(anomalies, key=lambda a: abs(a["z_score"]), reverse=True)


# ── expense projection (rolling avg + OLS trend on weekly card spend) ───────
# NOTE: this is EXPENSE-only. No income data exists in this repo, so unlike
# giovannini-finance's cash-flow projection, there is no balance/savings-rate
# figure here — only "expect to spend about X in the next N days".
def _ols_slope(y: list[float]) -> float:
    n = len(y)
    if n < 2:
        return 0.0
    x = list(range(n))
    x_mean = sum(x) / n
    y_mean = sum(y) / n
    num = sum((x[i] - x_mean) * (y[i] - y_mean) for i in range(n))
    den = sum((x[i] - x_mean) ** 2 for i in range(n))
    return num / den if den else 0.0


def project_expense(weekly: list[dict], horizons_days: tuple[int, ...] = (30, 60, 90)) -> dict:
    if len(weekly) < 3:
        return {"error": "not enough weekly records for a projection (need >= 3)"}

    spends = [w["spend"] for w in weekly]
    recent = spends[-4:] if len(spends) >= 4 else spends
    rolling_avg_weekly = statistics.mean(recent)
    slope_weekly = _ols_slope(spends[-8:] if len(spends) >= 8 else spends)

    projections = {}
    for days in horizons_days:
        weeks = days / 7
        trend_adjusted = rolling_avg_weekly + slope_weekly * (weeks / 2)
        projections[f"{days}d"] = round(max(trend_adjusted, 0) * weeks, 2)

    return {
        "rolling_avg_weekly": round(rolling_avg_weekly, 2),
        "trend_slope_weekly": round(slope_weekly, 2),
        "trend_direction": "subindo" if slope_weekly > 5 else "caindo" if slope_weekly < -5 else "estavel",
        "projected_spend": projections,
    }


# ── narrative (deterministic, PT-BR — no LLM, mirrors giovannini-finance) ───
def build_narrative(anomalies: list[dict], projection: dict) -> list[str]:
    bullets = []
    if anomalies:
        top = anomalies[0]
        bullets.append(
            f"Gasto em '{top['category']}' está {top['direction']} do padrão em "
            f"{top['month']} (R$ {top['current']:.2f} vs média de R$ {top['baseline_mean']:.2f}, "
            f"z={top['z_score']})."
        )
    else:
        bullets.append("Nenhuma anomalia de categoria AI/assinatura detectada no último mês.")

    if "projected_spend" in projection:
        d = projection["trend_direction"]
        bullets.append(
            f"Tendência de gasto semanal (cartão, todas categorias): {d}. "
            f"Projeção 30d: R$ {projection['projected_spend']['30d']:.2f}."
        )
    return bullets


# ── main ─────────────────────────────────────────────────────────────────────
def main() -> None:
    require_workbook()
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

    txs = load_transactions(wb)
    weekly = load_weekly(wb)
    monthly_cat = build_monthly_category_agg(txs)
    anomalies = detect_anomalies(monthly_cat)
    projection = project_expense(weekly)
    narrative = build_narrative(anomalies, projection)

    output = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "scope_note": (
            "Ported from giovannini-finance/scripts/compute_insights.py. "
            "Anomaly detection covers ONLY the AI/subscription Transactions "
            "subset (not full ledger). Projection is expense-only (no income "
            "data exists in this repo, so no cash-flow/balance figure)."
        ),
        "source": {
            "workbook": WORKBOOK_PATH.name,
            "transactions_count": len(txs),
            "weekly_records": len(weekly),
        },
        "anomalies": anomalies,
        "expense_projection": projection,
        "narrative": narrative,
    }

    OUTPUT_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH} ({len(anomalies)} anomalies, {len(narrative)} narrative bullets)")


if __name__ == "__main__":
    main()
